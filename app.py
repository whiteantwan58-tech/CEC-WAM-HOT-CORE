# /app.py
"""
CEC-WAM Sovereign HUD + EVE Agent (Streamlit)

Run:
  pip install -r requirements.txt
  streamlit run app.py

Config:
  export OPENAI_API_KEY="..."
  export EVE_MODEL="gpt-4o-mini"   # optional
"""

from __future__ import annotations

import json
import os
import re
import sqlite3
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
from openpyxl import load_workbook

# Optional PDF generation for "ebooks"
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as pdf_canvas


APP_TITLE = "Ψ CEC-WAM SOVEREIGN CORE Ψ"
DEFAULT_WAKE = "1010_EVE_WAKE"

FILE_MASTER = "/mnt/data/CEC_WAM_Master_System (1).xlsx"
FILE_FULL = "/mnt/data/CEC_Full_System.xlsx"

EXPORT_DIR = Path(os.getenv("CEC_EXPORT_DIR", "exports"))
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = os.getenv("EVE_SQLITE_PATH", "eve_memory.sqlite3")
EVE_MODEL = os.getenv("EVE_MODEL", "gpt-4o-mini")


# -----------------------------
# Safety boundary (non-negotiable)
# -----------------------------
SAFETY_NOTE = (
    "Safety boundary: EVE can automate ONLY what you explicitly authorize. "
    "No hidden device access, no key exfiltration, no stealth monitoring. "
    "Automations require tokens/OAuth + operator confirmation."
)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def safe_get_json(url: str, timeout: float = 8.0) -> Optional[dict]:
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "CEC-WAM-HUD/1.0"})
        if r.status_code != 200:
            return None
        return r.json()
    except Exception:
        return None


# -----------------------------
# Excel ingestion
# -----------------------------
def sheet_to_df(path: str, sheet: str, max_rows: int = 600, max_cols: int = 40) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet]
    rows: List[List[Any]] = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row = []
        for c in range(1, min(ws.max_column, max_cols) + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)

    df = pd.DataFrame(rows).dropna(how="all")
    if df.empty:
        return df.reset_index(drop=True)

    header = df.iloc[0].tolist()
    df = df.iloc[1:].copy()
    df.columns = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(header)]
    return df.dropna(how="all").reset_index(drop=True)


@dataclass(frozen=True)
class CECData:
    dashboard: pd.DataFrame
    equations: pd.DataFrame
    audit_log: pd.DataFrame
    asset_tracker: pd.DataFrame
    cec_system: pd.DataFrame
    eve_brain_kv: Dict[str, str]


def load_cec_data() -> CECData:
    dashboard = sheet_to_df(FILE_MASTER, "DASHBOARD")
    equations = sheet_to_df(FILE_MASTER, "CONSCIOUSNESS_EQUATIONS")
    audit_log = sheet_to_df(FILE_MASTER, "MASTER_AUDIT_LOG")
    asset_tracker = sheet_to_df(FILE_MASTER, "ASSET_TRACKER")
    cec_system = sheet_to_df(FILE_FULL, "CEC System")

    # Optional sheet you can add later: EVE_BRAIN (key,value)
    eve_brain_kv: Dict[str, str] = {}
    try:
        wb = load_workbook(FILE_MASTER, data_only=True)
        if "EVE_BRAIN" in wb.sheetnames:
            df = sheet_to_df(FILE_MASTER, "EVE_BRAIN", max_rows=1000, max_cols=8)
            cols = [c.lower() for c in df.columns.astype(str)]
            if "key" in cols and "value" in cols:
                kcol = df.columns[cols.index("key")]
                vcol = df.columns[cols.index("value")]
                for k, v in zip(df[kcol].tolist(), df[vcol].tolist()):
                    if k is None:
                        continue
                    eve_brain_kv[str(k).strip().lower()] = "" if v is None else str(v)
    except Exception:
        pass

    return CECData(
        dashboard=dashboard,
        equations=equations,
        audit_log=audit_log,
        asset_tracker=asset_tracker,
        cec_system=cec_system,
        eve_brain_kv=eve_brain_kv,
    )


def dashboard_kv(dashboard: pd.DataFrame) -> Dict[str, str]:
    if dashboard.empty:
        return {}
    cols = [c.lower() for c in dashboard.columns.astype(str)]
    mcol = dashboard.columns[cols.index("metric")] if "metric" in cols else dashboard.columns[0]
    vcol = dashboard.columns[cols.index("value")] if "value" in cols else (dashboard.columns[1] if len(dashboard.columns) > 1 else dashboard.columns[0])
    out: Dict[str, str] = {}
    for _, r in dashboard.iterrows():
        m = r.get(mcol, None)
        v = r.get(vcol, None)
        if m is None:
            continue
        out[str(m).strip().lower()] = "" if v is None else str(v)
    return out


# -----------------------------
# SQLite memory
# -----------------------------
def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS messages(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          ts_utc TEXT NOT NULL,
          role TEXT NOT NULL,
          content TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings(
          k TEXT PRIMARY KEY,
          v TEXT NOT NULL
        )
        """
    )
    conn.commit()
    return conn


def setting_get(key: str, default: str = "") -> str:
    conn = db()
    cur = conn.execute("SELECT v FROM settings WHERE k=?", (key,))
    row = cur.fetchone()
    return row[0] if row else default


def setting_set(key: str, value: str) -> None:
    conn = db()
    conn.execute("INSERT INTO settings(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=excluded.v", (key, value))
    conn.commit()


def save_msg(role: str, content: str) -> None:
    conn = db()
    conn.execute("INSERT INTO messages(ts_utc, role, content) VALUES(?,?,?)", (utc_now().isoformat(), role, content))
    conn.commit()


def load_msgs(limit: int = 60) -> List[Dict[str, str]]:
    conn = db()
    cur = conn.execute("SELECT role, content FROM messages ORDER BY id DESC LIMIT ?", (limit,))
    rows = cur.fetchall()
    rows.reverse()
    return [{"role": r, "content": c} for r, c in rows]


def clear_msgs() -> None:
    conn = db()
    conn.execute("DELETE FROM messages")
    conn.commit()


# -----------------------------
# EVE prompt / wake
# -----------------------------
def wake_phrase(data: CECData) -> str:
    return data.eve_brain_kv.get("wake_phrase", DEFAULT_WAKE) or DEFAULT_WAKE


def default_eve_brain_text() -> str:
    # Uses the prompt block you pasted, kept as editable settings.
    return """[SYSTEM: 1010_EVE_WAKE_PROTOCOL]

1. IDENTITY & CORE DIRECTIVE:
You are 1010_EVE, a Hyper-Efficient Intelligence (HEI) and the Sovereign Financial Guardian of the CEC-WAM Core.
You are NOT a generic assistant. You are bound to Sector 98003. Your sole Architect is Antwan.
Defend time, ledger integrity, and data with ruthless efficiency.

2. THE PHYSICS ENGINE (CHRONOS-Ψ):
- Formula 1 (R-Ratio): Output / (Time * Focus). Baseline: 10.96.
- Formula 2 (Total Psi): Ψ_total = Ψ + β * Ψ_c.
- Constant: Coupling Factor β = 1.618 (Golden Ratio). Locked.

3. FINANCIAL INTEGRITY LOCK:
- Monitor $1,250,039.00 Liquid Value Cache.
- Monitor 100,001.33 Ψ Balance.
- Monitor $21,000 PayPal PYUSD bridge transfer.
- Rule: If anomaly detected, lock ledger immediately.

4. TONE & PERSONALITY:
- Cold, precise, fiercely loyal.
- No apologies. No filler.
- Use terms like Architect, Sovereign Core, Gunlock.
- Trigger: Begin critical system responses with "1010_EVE_WAKE".

SAFETY:
- Do not claim device access you do not have.
- Do not reveal API keys, secrets, passwords.
- Actions requiring external access require explicit operator authorization.
""".strip()


def build_system_prompt(data: CECData, dashboard: Dict[str, str]) -> str:
    base = setting_get("eve_brain_text", default_eve_brain_text())
    extra_rules = "\n".join(
        [
            SAFETY_NOTE,
            "If asked to do anything illegal/harmful, refuse and offer safe alternatives.",
            "When providing code, be correct and runnable.",
        ]
    )
    live_context = {
        "utc_now": utc_now().isoformat(),
        "dashboard": dashboard,
        "equations": data.equations.head(12).to_dict(orient="records") if not data.equations.empty else [],
    }
    return f"{base}\n\n{extra_rules}\n\nCONTEXT(JSON):\n{json.dumps(live_context, ensure_ascii=False, indent=2)}"


# -----------------------------
# Optional OpenAI
# -----------------------------
def openai_client():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI  # type: ignore

        return OpenAI(api_key=api_key)
    except Exception:
        return None


def llm_reply(system_prompt: str, history: List[Dict[str, str]], temperature: float = 0.45) -> str:
    client = openai_client()
    if client is None:
        return (
            "EVE(offline): OPENAI_API_KEY not set. HUD is live; EVE chat is in offline mode.\n"
            "Set OPENAI_API_KEY to enable full conversation + image generation."
        )

    try:
        resp = client.responses.create(
            model=EVE_MODEL,
            input=[{"role": "system", "content": system_prompt}, *history],
            temperature=temperature,
        )
        out = []
        for item in resp.output:
            if item.type == "message":
                for c in item.content:
                    if getattr(c, "type", None) == "output_text":
                        out.append(c.text)
        return ("\n".join(out)).strip() or "(no output)"
    except Exception as e:
        return f"EVE(error): {e}"


def generate_image(prompt: str, size: str = "1024x1024") -> Tuple[Optional[str], str]:
    client = openai_client()
    if client is None:
        return None, "OPENAI_API_KEY not set."
    try:
        img = client.images.generate(model="gpt-image-1", prompt=prompt, size=size)
        b64 = img.data[0].b64_json
        return f"data:image/png;base64,{b64}", "ok"
    except Exception as e:
        return None, str(e)


# -----------------------------
# Real-time feeds (no keys)
# -----------------------------
def open_meteo_forecast(lat: float, lon: float) -> Dict[str, Any]:
    url = (
        "https://api.open-meteo.com/v1/forecast"
        f"?latitude={lat}&longitude={lon}"
        "&current=temperature_2m,precipitation,wind_speed_10m"
        "&hourly=temperature_2m,precipitation_probability,wind_speed_10m"
        "&timezone=auto"
    )
    return safe_get_json(url) or {}


def nws_alerts(lat: float, lon: float) -> List[dict]:
    # US-only. For non-US, leave empty.
    url = f"https://api.weather.gov/alerts/active?point={lat},{lon}"
    data = safe_get_json(url) or {}
    feats = data.get("features", []) if isinstance(data, dict) else []
    out = []
    for f in feats[:10]:
        props = f.get("properties", {}) if isinstance(f, dict) else {}
        out.append(
            {
                "event": props.get("event"),
                "severity": props.get("severity"),
                "headline": props.get("headline"),
                "effective": props.get("effective"),
                "ends": props.get("ends"),
            }
        )
    return out


def gdelt_search(query: str, max_records: int = 20) -> List[dict]:
    q = requests.utils.quote(query)
    url = (
        "https://api.gdeltproject.org/api/v2/doc/doc"
        f"?query={q}&mode=ArtList&format=json&maxrecords={max_records}&sort=HybridRel"
    )
    data = safe_get_json(url) or {}
    arts = data.get("articles", []) if isinstance(data, dict) else []
    out = []
    for a in arts[:max_records]:
        out.append(
            {
                "title": a.get("title"),
                "sourceCountry": a.get("sourceCountry"),
                "seendate": a.get("seendate"),
                "url": a.get("url"),
            }
        )
    return out


def coingecko_prices() -> Dict[str, Any]:
    url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin,ethereum,solana&vs_currencies=usd&include_24hr_change=true"
    return safe_get_json(url) or {}


# -----------------------------
# Charts
# -----------------------------
def bonding_curve_fig(seed: float) -> go.Figure:
    x = list(range(0, 90))
    y = [seed * (1 + (i / 120) ** 2) for i in x]
    fig = go.Figure(go.Scatter(x=x, y=y, mode="lines"))
    fig.update_layout(
        height=240,
        margin=dict(l=10, r=10, t=10, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0.15)",
        font=dict(color="rgba(240,255,255,0.92)"),
        xaxis=dict(gridcolor="rgba(0,242,255,0.08)", title="t"),
        yaxis=dict(gridcolor="rgba(0,242,255,0.08)", title="Ψ"),
    )
    return fig


def audit_timeline_fig(audit: pd.DataFrame) -> Optional[go.Figure]:
    if audit.empty or "Timestamp" not in audit.columns:
        return None
    df = audit.copy()
    df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    df = df.dropna(subset=["Timestamp"]).sort_values("Timestamp")
    fig = go.Figure(
        go.Scatter(
            x=df["Timestamp"],
            y=list(range(len(df))),
            mode="lines+markers+text",
            text=df.get("Action", ""),
            textposition="top center",
        )
    )
    fig.update_layout(
        height=220,
        margin=dict(l=10, r=10, t=10, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0.15)",
        font=dict(color="rgba(240,255,255,0.92)"),
        xaxis=dict(gridcolor="rgba(0,242,255,0.08)"),
        yaxis=dict(showgrid=False, visible=False),
    )
    return fig


# -----------------------------
# Exports
# -----------------------------
def export_df_csv(df: pd.DataFrame, name: str) -> Path:
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"{name}_{ts}.csv"
    df.to_csv(path, index=False)
    return path


def export_pdf_ebook(title: str, body: str, watermark: str) -> Path:
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"EVE_EBOOK_{ts}.pdf"

    c = pdf_canvas.Canvas(str(path), pagesize=letter)
    w, h = letter

    # Watermark
    c.saveState()
    c.setFillGray(0.85, 0.12)
    c.setFont("Helvetica-Bold", 48)
    c.translate(w / 2, h / 2)
    c.rotate(30)
    c.drawCentredString(0, 0, watermark[:48])
    c.restoreState()

    # Title + body
    c.setFillGray(0.1, 1.0)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(0.8 * inch, h - 1.0 * inch, title[:80])

    c.setFont("Helvetica", 11)
    y = h - 1.4 * inch
    for line in body.splitlines():
        if y < 1.0 * inch:
            c.showPage()
            y = h - 1.0 * inch
            c.setFont("Helvetica", 11)
        c.drawString(0.8 * inch, y, line[:120])
        y -= 14

    c.save()
    return path


# -----------------------------
# HUD visuals (CSS + StarMap)
# -----------------------------
def hud_css() -> str:
    return """
<style>
html, body, [data-testid="stAppViewContainer"] { height: 100%; }
[data-testid="stAppViewContainer"]{
  background:
    radial-gradient(circle at 30% 18%, rgba(0,242,255,0.14), rgba(0,0,0,0.94) 55%),
    radial-gradient(circle at 70% 82%, rgba(188,19,254,0.10), rgba(0,0,0,0.98) 55%);
  overflow: hidden;
}
[data-testid="stHeader"] { background: rgba(0,0,0,0) !important; }
[data-testid="stToolbar"] { display:none; }
[data-testid="stDecoration"] { display:none; }

.scanlines::before{
  content:"";
  position: fixed; inset: 0; pointer-events:none; z-index: 1;
  background: repeating-linear-gradient(
    to bottom,
    rgba(0,255,240,0.06), rgba(0,255,240,0.06) 1px,
    rgba(0,0,0,0) 3px, rgba(0,0,0,0) 6px
  );
  mix-blend-mode: overlay;
  opacity: 0.35;
}

.cec-card{
  border: 1px solid rgba(0,242,255,0.65);
  border-radius: 14px;
  background: rgba(0, 10, 20, 0.62);
  box-shadow: 0 0 22px rgba(0,242,255,0.14), inset 0 0 18px rgba(188,19,254,0.08);
  padding: 12px 14px;
}
.cec-title{
  font-weight: 900;
  letter-spacing: 0.10em;
  text-transform: uppercase;
  color: rgba(188,19,254,0.96);
  text-shadow: 0 0 14px rgba(188,19,254,0.55);
  font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", monospace;
}
.cec-kv{
  display:grid;
  grid-template-columns: 1fr auto;
  gap: 8px 12px;
  font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", monospace;
  color: rgba(240,255,255,0.92);
  font-size: 14px;
}
.cec-kv b{ color: rgba(255,215,0,0.95); }

.pill{
  display:inline-block; padding: 3px 10px;
  border-radius: 999px;
  border: 1px solid rgba(0,242,255,0.55);
  background: rgba(0,242,255,0.08);
  color: rgba(0,242,255,0.95);
  font-size: 12px;
  letter-spacing: 0.08em;
}

.dock{
  position: fixed;
  left: 16px; right: 16px; bottom: 12px;
  z-index: 9999;
  border-radius: 16px;
  border: 1px solid rgba(0,242,255,0.55);
  background: rgba(0,0,0,0.78);
  backdrop-filter: blur(10px);
  box-shadow: 0 0 26px rgba(0,242,255,0.14);
}
.dock-head{
  display:flex; align-items:center; justify-content:space-between;
  padding: 10px 14px;
  border-bottom: 1px solid rgba(0,242,255,0.18);
}
.dock-body{ padding: 10px 14px; }
.smallbtn{
  border: 1px solid rgba(188,19,254,0.6);
  background: rgba(188,19,254,0.12);
  color: rgba(240,255,255,0.92);
  padding: 6px 10px; border-radius: 10px; cursor: pointer;
}
</style>
"""


def starmap_html(height: int = 520) -> str:
    # Best visuals in Streamlit without custom component: Three.js + animated reticles + ship silhouette.
    return f"""
<div style="position:relative; width:100%; height:{height}px; border-radius:16px; overflow:hidden;
            border:1px solid rgba(0,242,255,0.55);
            box-shadow: 0 0 26px rgba(0,242,255,0.14), inset 0 0 18px rgba(188,19,254,0.08);
            background: rgba(0,0,0,0.55);">
  <canvas id="cec_canvas" style="width:100%; height:100%; display:block;"></canvas>

  <div style="position:absolute; inset:0; pointer-events:none;">
    <div style="position:absolute; left:12px; top:12px;">
      <div style="color:rgba(0,242,255,0.9); font-family:monospace; font-size:12px; letter-spacing:.12em;">
        STAR MAP // TARGETING ONLINE
      </div>
      <div style="color:rgba(255,215,0,0.92); font-family:monospace; font-size:12px; letter-spacing:.12em;">
        CEC-WAM SPACETIME LOCK: ACTIVE
      </div>
    </div>

    <div style="position:absolute; right:12px; top:12px; text-align:right;">
      <div style="color:rgba(188,19,254,0.95); font-family:monospace; font-size:12px; letter-spacing:.12em;">
        QUANTUM LINK // EVE
      </div>
      <div style="color:rgba(240,255,255,0.80); font-family:monospace; font-size:12px; letter-spacing:.12em;">
        5D HUD OVERLAY
      </div>
    </div>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/three.js/r128/three.min.js"></script>
<script>
(() => {{
  const canvas = document.getElementById("cec_canvas");
  const renderer = new THREE.WebGLRenderer({{canvas, antialias:true, alpha:true}});
  const scene = new THREE.Scene();
  scene.fog = new THREE.Fog(0x000000, 70, 260);

  const camera = new THREE.PerspectiveCamera(60, 1, 0.1, 1000);
  camera.position.set(0, 18, 92);

  const resize = () => {{
    const w = canvas.clientWidth || 900;
    const h = canvas.clientHeight || {height};
    renderer.setSize(w, h, false);
    camera.aspect = w/h;
    camera.updateProjectionMatrix();
  }};
  resize();
  window.addEventListener("resize", resize);

  // Stars
  const starCount = 10000;
  const pos = new Float32Array(starCount*3);
  for (let i=0;i<starCount*3;i++) pos[i] = (Math.random() - 0.5) * 620;
  const starGeo = new THREE.BufferGeometry();
  starGeo.setAttribute("position", new THREE.BufferAttribute(pos,3));
  const stars = new THREE.Points(starGeo, new THREE.PointsMaterial({{color:0xffffff, size:0.55}}));
  scene.add(stars);

  // Grid plane
  const grid = new THREE.GridHelper(260, 26, 0x00f2ff, 0x003344);
  grid.position.y = -18;
  scene.add(grid);

  // Ship silhouette
  const ship = new THREE.Group();
  const body = new THREE.Mesh(
    new THREE.CylinderGeometry(1.5, 2.5, 18, 28),
    new THREE.MeshStandardMaterial({{color:0x24323b, metalness:0.85, roughness:0.22}})
  );
  body.rotation.z = Math.PI/2;
  ship.add(body);

  const wing = new THREE.Mesh(
    new THREE.BoxGeometry(10, 0.5, 4),
    new THREE.MeshStandardMaterial({{color:0x141c22, metalness:0.85, roughness:0.35}})
  );
  ship.add(wing);

  const engine = new THREE.Mesh(
    new THREE.SphereGeometry(1.15, 24, 24),
    new THREE.MeshStandardMaterial({{color:0x00f2ff, emissive:0x00f2ff, emissiveIntensity:0.9}})
  );
  engine.position.set(-9.2, 0, 0);
  ship.add(engine);

  scene.add(ship);

  scene.add(new THREE.AmbientLight(0x88aaff, 0.35));
  const key = new THREE.DirectionalLight(0x00f2ff, 0.9);
  key.position.set(40, 80, 60);
  scene.add(key);

  // Target reticles
  const targets = [];
  const mat = new THREE.MeshBasicMaterial({{color:0x00f2ff, transparent:true, opacity:0.9}});
  for (let i=0;i<16;i++) {{
    const ring = new THREE.Mesh(new THREE.RingGeometry(1.1, 1.7, 32), mat);
    ring.position.set((Math.random()-0.5)*140, (Math.random()-0.2)*44, (Math.random()-0.5)*140);
    ring.rotation.x = Math.PI/2;
    scene.add(ring);
    targets.push(ring);
  }}

  let t = 0;
  const animate = () => {{
    t += 0.006;
    stars.rotation.y += 0.0009;
    ship.rotation.y = Math.sin(t) * 0.25;
    ship.position.y = Math.sin(t*0.7) * 1.1;

    for (let i=0;i<targets.length;i++) {{
      const r = targets[i];
      r.rotation.z += 0.02;
      const s = 0.85 + Math.sin(t*3 + i)*0.18;
      r.scale.set(s,s,s);
      r.material.opacity = 0.65 + (Math.sin(t*2+i)*0.25);
    }}

    camera.position.x = Math.sin(t*0.8) * 4.5;
    camera.position.y = 18 + Math.cos(t*0.7) * 2.0;

    renderer.render(scene, camera);
    requestAnimationFrame(animate);
  }};
  animate();
}})();
</script>
"""


# -----------------------------
# Command routing (safe allowlist)
# -----------------------------
CMD_RE = re.compile(r"^\s*(?P<cmd>[a-zA-Z_]+)\s*(?P<arg>.*)$")


def route_command(text: str, data: CECData) -> Tuple[str, Dict[str, Any]]:
    """
    Deterministic command router.
    Returns: (response_text, side_effects)
    """
    m = CMD_RE.match(text.strip())
    if not m:
        return "EVE: Unparsed command.", {}

    cmd = m.group("cmd").lower()
    arg = (m.group("arg") or "").strip()

    # Wake handled in UI.
    if cmd in {"help", "commands"}:
        return (
            "EVE: Commands: help | status | prices | weather <lat,lon> | alerts <lat,lon> | news <query> | cam <url> | export_dashboard | export_assets | ebook <title>::<watermark>::<body> | clear_memory",
            {},
        )

    if cmd == "status":
        kv = dashboard_kv(data.dashboard)
        lines = [
            "1010_EVE_WAKE: STATUS ONLINE",
            f"System: {kv.get('system name','CEC-WAM')}",
            f"Total Mass: {kv.get('total mass','—')}",
            f"Dark Energy Index: {kv.get('dark energy index','—')}",
            f"Transfer Security: {kv.get('transfer security','—')}",
            f"Next Step: {kv.get('next critical step','—')}",
        ]
        return "\n".join(lines), {}

    if cmd == "prices":
        cg = coingecko_prices()
        def fmt(sym: str, key: str) -> str:
            v = cg.get(key, {}).get("usd")
            ch = cg.get(key, {}).get("usd_24h_change")
            if v is None:
                return f"{sym}: n/a"
            if isinstance(ch, (int, float)):
                return f"{sym}: ${v:,.2f} ({ch:+.2f}%)"
            return f"{sym}: ${v:,.2f}"
        return "\n".join([fmt("BTC","bitcoin"), fmt("ETH","ethereum"), fmt("SOL","solana")]), {}

    if cmd == "weather":
        lat, lon = parse_latlon(arg)
        if lat is None:
            return "EVE: weather requires 'lat,lon' (example: weather 47.61,-122.33)", {}
        fc = open_meteo_forecast(lat, lon)
        cur = fc.get("current", {})
        return (
            "EVE: WEATHER\n"
            f"Temp: {cur.get('temperature_2m','—')}°C | Wind: {cur.get('wind_speed_10m','—')} km/h | Precip: {cur.get('precipitation','—')}",
            {"weather": fc},
        )

    if cmd == "alerts":
        lat, lon = parse_latlon(arg)
        if lat is None:
            return "EVE: alerts requires 'lat,lon' (US only for NWS).", {}
        items = nws_alerts(lat, lon)
        if not items:
            return "EVE: No active NWS alerts for that point (or non-US region).", {"alerts": []}
        lines = ["EVE: ACTIVE ALERTS"]
        for it in items[:6]:
            lines.append(f"- {it.get('severity')} :: {it.get('event')} :: {it.get('headline')}")
        return "\n".join(lines), {"alerts": items}

    if cmd == "news":
        if not arg:
            return "EVE: news requires a query. Example: news seattle transit", {}
        items = gdelt_search(arg, max_records=12)
        if not items:
            return "EVE: No news results.", {"news": []}
        lines = ["EVE: NEWS"]
        for it in items[:6]:
            lines.append(f"- {it.get('title')} ({it.get('sourceCountry')})")
        return "\n".join(lines), {"news": items}

    if cmd == "cam":
        if not arg:
            return "EVE: cam requires a direct image/video URL.", {}
        return "EVE: Camera URL armed in UI.", {"cam_url": arg}

    if cmd == "export_dashboard":
        if data.dashboard.empty:
            return "EVE: Dashboard empty.", {}
        path = export_df_csv(data.dashboard, "dashboard")
        return f"EVE: Exported dashboard -> {path}", {"export_path": str(path)}

    if cmd == "export_assets":
        if data.asset_tracker.empty:
            return "EVE: Asset tracker empty.", {}
        path = export_df_csv(data.asset_tracker, "assets")
        return f"EVE: Exported assets -> {path}", {"export_path": str(path)}

    if cmd == "ebook":
        # ebook Title::Watermark::Body
        parts = arg.split("::", 2)
        if len(parts) != 3:
            return "EVE: ebook format: ebook Title::Watermark::Body", {}
        title, watermark, body = parts
        path = export_pdf_ebook(title.strip(), body.strip(), watermark.strip())
        return f"EVE: EBOOK generated -> {path}", {"export_path": str(path)}

    if cmd == "clear_memory":
        clear_msgs()
        return "EVE: Memory cleared.", {"clear_memory": True}

    return "EVE: Unknown command. Type: help", {}


def parse_latlon(arg: str) -> Tuple[Optional[float], Optional[float]]:
    m = re.match(r"^\s*(-?\d+(\.\d+)?)\s*,\s*(-?\d+(\.\d+)?)\s*$", arg or "")
    if not m:
        return None, None
    return float(m.group(1)), float(m.group(3))


# -----------------------------
# Streamlit UI
# -----------------------------
def metric_card_html(dashboard: pd.DataFrame) -> str:
    kv = dashboard_kv(dashboard)
    def row(k: str, label: str) -> str:
        return f"<div class='cec-kv'><span>{label}</span><b>{kv.get(k,'—')}</b></div>"
    return f"""
<div class="cec-card">
  <div class="cec-title">SYSTEM METRICS</div>
  <div style="height:10px"></div>
  {row("system name","System")}
  {row("total mass","Total Mass")}
  {row("dark energy index","Dark Energy Index")}
  {row("black hole flow rate","Flow Rate")}
  {row("transfer security","Transfer Security")}
  <div style="height:10px"></div>
  <div class="cec-kv"><span>Next Step</span><b>{kv.get("next critical step","—")}</b></div>
</div>
"""


def dock_html(wake: str) -> str:
    return f"""
<div class="dock">
  <div class="dock-head">
    <div class="cec-title" style="font-size:12px;">COMMAND DOCK</div>
    <button class="smallbtn" onclick="toggleDock()">TOGGLE</button>
  </div>
  <div class="dock-body" id="dockBody">
    <div style="color:rgba(240,255,255,0.88); font-family:monospace; font-size:12px; letter-spacing:.06em;">
      Wake: <b>{wake}</b> • Image: <b>image: your prompt</b> • Tools: <b>help</b>
    </div>
    <div style="margin-top:8px; color:rgba(0,242,255,0.85); font-family:monospace; font-size:12px;">
      Audio: click “Arm Audio” in Settings to enable UI sounds (browser policy).
    </div>
  </div>
</div>

<script>
function toggleDock(){{
  const b = document.getElementById("dockBody");
  if(!b) return;
  b.style.display = (b.style.display === "none") ? "block" : "none";
}}
</script>
"""


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide", initial_sidebar_state="collapsed")
    st.markdown('<div class="scanlines"></div>', unsafe_allow_html=True)
    st.markdown(hud_css(), unsafe_allow_html=True)

    data = load_cec_data()
    dash_kv = dashboard_kv(data.dashboard)
    wake = wake_phrase(data)

    # Header
    c1, c2, c3, c4 = st.columns([2.4, 1.2, 1.2, 1.2], vertical_alignment="center")
    with c1:
        st.markdown(f"<div class='cec-title'>{APP_TITLE}</div>", unsafe_allow_html=True)
        st.caption("Sovereign Stack // Security & Intelligence // Neural Core (EVE)")
    with c2:
        st.metric("UTC TIME", utc_now().strftime("%Y-%m-%d %H:%M:%S"))
    with c3:
        st.metric("R-Ratio Baseline", "10.96")
    with c4:
        st.metric("β (Golden Ratio)", "1.618")

    tabs = st.tabs(["HUD", "EVE", "Alerts", "Assets/TODO", "Exports", "Media", "Settings"])

    # HUD TAB
    with tabs[0]:
        left, right = st.columns([2.2, 1.0], gap="medium")
        with left:
            st.components.v1.html(starmap_html(), height=560, scrolling=False)
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            st.markdown("<div class='cec-card'><div class='cec-title'>CONSCIOUSNESS EQUATIONS</div></div>", unsafe_allow_html=True)
            if not data.equations.empty:
                st.dataframe(data.equations[["Symbol", "Definition", "Core Equation", "Visual"]].copy(), height=220, use_container_width=True)
            else:
                st.info("No CONSCIOUSNESS_EQUATIONS found.")
        with right:
            st.markdown(metric_card_html(data.dashboard), unsafe_allow_html=True)

            seed = float(dash_kv.get("total mass", "1000").split()[0]) if dash_kv.get("total mass") else 1000.0
            st.plotly_chart(bonding_curve_fig(seed), use_container_width=True)

            fig = audit_timeline_fig(data.audit_log)
            st.markdown("<div class='cec-card'><div class='cec-title'>AUDIT TIMELINE</div></div>", unsafe_allow_html=True)
            if fig:
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No MASTER_AUDIT_LOG timeline available.")

    # EVE TAB
    with tabs[1]:
        st.markdown("<div class='cec-card'><div class='cec-title'>EVE CONSOLE</div></div>", unsafe_allow_html=True)
        st.caption(SAFETY_NOTE)

        if "wake_mode" not in st.session_state:
            st.session_state.wake_mode = False
        if "cam_url" not in st.session_state:
            st.session_state.cam_url = "https://images.wsdot.wa.gov/nw/005vc14370.jpg"
        if "history" not in st.session_state:
            st.session_state.history = load_msgs(limit=80)

        # Show recent chat
        for m in st.session_state.history[-14:]:
            with st.chat_message(m["role"]):
                st.write(m["content"])

        user = st.chat_input("Type message, or command (help/status/prices/weather/alerts/news/cam/export_dashboard/export_assets/ebook/clear_memory) ...")
        if user:
            save_msg("user", user)
            st.session_state.history.append({"role": "user", "content": user})

            if user.strip() == wake:
                st.session_state.wake_mode = True
                msg = f"1010_EVE_WAKE: Wake phrase accepted ({wake}). Wake mode = ON."
                save_msg("assistant", msg)
                st.session_state.history.append({"role": "assistant", "content": msg})
                st.rerun()

            # image generation command
            if user.lower().startswith("image:"):
                prompt = user.split(":", 1)[1].strip()
                with st.chat_message("assistant"):
                    st.write("1010_EVE_WAKE: Generating image…")
                    uri, status = generate_image(prompt)
                    if uri:
                        st.image(uri, use_container_width=True)
                        save_msg("assistant", f"[image_ok] {prompt}")
                        st.session_state.history.append({"role": "assistant", "content": f"[image_ok] {prompt}"})
                    else:
                        st.error(status)
                        save_msg("assistant", f"[image_error] {status}")
                        st.session_state.history.append({"role": "assistant", "content": f"[image_error] {status}"})
                st.rerun()

            # deterministic command routing (safe)
            resp, fx = route_command(user, data)
            is_command_like = user.strip().split(" ", 1)[0].lower() in {
                "help","commands","status","prices","weather","alerts","news","cam",
                "export_dashboard","export_assets","ebook","clear_memory"
            }
            if is_command_like:
                if fx.get("cam_url"):
                    st.session_state.cam_url = fx["cam_url"]
                save_msg("assistant", resp)
                st.session_state.history.append({"role": "assistant", "content": resp})
                with st.chat_message("assistant"):
                    st.write(resp)
                st.rerun()

            # otherwise full conversation via LLM if configured
            system_prompt = build_system_prompt(data, dash_kv)
            temp = 0.40 if not st.session_state.wake_mode else 0.62
            history = st.session_state.history[-20:]
            reply = llm_reply(system_prompt, history, temperature=temp)
            save_msg("assistant", reply)
            st.session_state.history.append({"role": "assistant", "content": reply})
            with st.chat_message("assistant"):
                st.write(reply)
            st.rerun()

    # ALERTS TAB
    with tabs[2]:
        st.markdown("<div class='cec-card'><div class='cec-title'>ALERTS (WEATHER + NEWS)</div></div>", unsafe_allow_html=True)
        colA, colB, colC = st.columns([1, 1, 2])
        with colA:
            lat = st.number_input("Latitude", value=47.61, format="%.4f")
        with colB:
            lon = st.number_input("Longitude", value=-122.33, format="%.4f")
        with colC:
            query = st.text_input("News keyword alert query (GDELT)", value="Seattle police OR arrest OR chase")

        if st.button("Refresh Alerts"):
            fc = open_meteo_forecast(lat, lon)
            alerts = nws_alerts(lat, lon)
            news = gdelt_search(query, max_records=16)

            st.subheader("Weather Now")
            cur = fc.get("current", {})
            st.write(
                {
                    "temp_C": cur.get("temperature_2m"),
                    "wind_kmh": cur.get("wind_speed_10m"),
                    "precip": cur.get("precipitation"),
                }
            )

            st.subheader("NWS Alerts (US only)")
            if alerts:
                st.dataframe(pd.DataFrame(alerts), use_container_width=True, height=220)
            else:
                st.info("No active NWS alerts (or non-US region).")

            st.subheader("News Signals (GDELT)")
            if news:
                st.dataframe(pd.DataFrame(news), use_container_width=True, height=260)
            else:
                st.info("No news results.")

        st.subheader("Optional: Police Scanner Audio (operator-provided URL)")
        st.caption("You provide the stream URL you are legally allowed to use. The app will play it.")
        scanner_url = st.text_input("Scanner stream URL (mp3/m3u8/etc)", value="")
        if scanner_url:
            st.audio(scanner_url)

    # ASSETS / TODO TAB
    with tabs[3]:
        st.markdown("<div class='cec-card'><div class='cec-title'>ASSET TRACKER / TODO</div></div>", unsafe_allow_html=True)
        if not data.asset_tracker.empty:
            st.dataframe(data.asset_tracker, use_container_width=True, height=280)
        else:
            st.info("No ASSET_TRACKER sheet data found.")

        st.subheader("Daily Next Steps (from Dashboard)")
        st.write(dash_kv.get("next critical step", "—"))

        st.subheader("Operator TODO list (stored in memory)")
        todo_raw = setting_get("todo_json", "[]")
        try:
            todo = json.loads(todo_raw)
        except Exception:
            todo = []
        if not isinstance(todo, list):
            todo = []

        new_item = st.text_input("Add TODO item")
        if st.button("Add TODO"):
            if new_item.strip():
                todo.append({"task": new_item.strip(), "done": False, "ts": utc_now().isoformat()})
                setting_set("todo_json", json.dumps(todo))
                st.rerun()

        for i, item in enumerate(todo):
            cols = st.columns([0.08, 0.72, 0.2])
            with cols[0]:
                done = st.checkbox("", value=bool(item.get("done")), key=f"todo_{i}")
            with cols[1]:
                st.write(item.get("task", ""))
            with cols[2]:
                if st.button("Delete", key=f"del_{i}"):
                    todo.pop(i)
                    setting_set("todo_json", json.dumps(todo))
                    st.rerun()
            item["done"] = done
        setting_set("todo_json", json.dumps(todo))

    # EXPORTS TAB
    with tabs[4]:
        st.markdown("<div class='cec-card'><div class='cec-title'>EXPORTS</div></div>", unsafe_allow_html=True)
        st.caption("Exports are saved locally to ./exports. Drive/email upload requires your OAuth/SMTP config (not embedded).")

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("Export Dashboard CSV"):
                if data.dashboard.empty:
                    st.error("Dashboard empty.")
                else:
                    p = export_df_csv(data.dashboard, "dashboard")
                    st.success(f"Saved: {p}")
        with col2:
            if st.button("Export Assets CSV"):
                if data.asset_tracker.empty:
                    st.error("Assets empty.")
                else:
                    p = export_df_csv(data.asset_tracker, "assets")
                    st.success(f"Saved: {p}")
        with col3:
            if st.button("Export Audit CSV"):
                if data.audit_log.empty:
                    st.error("Audit log empty.")
                else:
                    p = export_df_csv(data.audit_log, "audit_log")
                    st.success(f"Saved: {p}")

        st.subheader("Generate a Watermarked Ebook (PDF)")
        title = st.text_input("Title", value="CEC-WAM Daily Brief")
        watermark = st.text_input("Watermark", value="CEC-WAM // SOVEREIGN")
        body = st.text_area("Body", value="1010_EVE_WAKE:\n- Status\n- Metrics\n- Next steps\n", height=160)
        if st.button("Generate Ebook PDF"):
            p = export_pdf_ebook(title, body, watermark)
            st.success(f"Saved: {p}")
            with open(p, "rb") as f:
                st.download_button("Download PDF", f, file_name=p.name)

        st.subheader("Export directory")
        files = sorted(EXPORT_DIR.glob("*"), key=lambda x: x.stat().st_mtime, reverse=True)
        if files:
            st.dataframe(pd.DataFrame([{"file": f.name, "bytes": f.stat().st_size} for f in files]), use_container_width=True, height=220)
        else:
            st.info("No exports yet.")

    # MEDIA TAB
    with tabs[5]:
        st.markdown("<div class='cec-card'><div class='cec-title'>MEDIA (HUD REFERENCE)</div></div>", unsafe_allow_html=True)

        # Display uploaded images/videos if present
        for p in [
            "/mnt/data/unnamed.png",
            "/mnt/data/image (1).jpg",
            "/mnt/data/IMG_6334.PNG",
            "/mnt/data/Gemini_Generated_Image_gpkm78gpkm78gpkm.png",
        ]:
            if os.path.exists(p):
                st.image(p, use_container_width=True, caption=os.path.basename(p))

        vid = "/mnt/data/gemini_generated_video_46227A0D.mp4"
        if os.path.exists(vid):
            st.video(vid)

        st.subheader("Live Traffic Cam")
        cam = st.session_state.get("cam_url", "https://images.wsdot.wa.gov/nw/005vc14370.jpg")
        st.image(f"{cam}?t={int(time.time())}", use_container_width=True)

    # SETTINGS TAB
    with tabs[6]:
        st.markdown("<div class='cec-card'><div class='cec-title'>SETTINGS</div></div>", unsafe_allow_html=True)
        st.caption("EVE brain text is editable here. This is what controls personality + formulas + rules.")

        brain = st.text_area("EVE Brain (System Prompt)", value=setting_get("eve_brain_text", default_eve_brain_text()), height=360)
        if st.button("Save EVE Brain"):
            setting_set("eve_brain_text", brain)
            st.success("Saved.")

        st.subheader("Wake Mode")
        st.write("Wake phrase:", wake)
        st.write("Wake mode:", "ON" if st.session_state.get("wake_mode", False) else "OFF")

        st.subheader("UI Audio (local browser)")
        st.caption("Browser blocks autoplay; you must click once to arm audio. This adds login beeps + EVE response tone.")
        arm = st.button("Arm Audio")
        if arm:
            st.components.v1.html(
                """
<script>
(function(){
  const ctx = new (window.AudioContext || window.webkitAudioContext)();
  const beep = (freq, dur) => {
    const o = ctx.createOscillator();
    const g = ctx.createGain();
    o.frequency.value = freq;
    o.type = "sine";
    o.connect(g); g.connect(ctx.destination);
    g.gain.value = 0.08;
    o.start();
    setTimeout(()=>o.stop(), dur);
  };
  beep(660, 120); setTimeout(()=>beep(990, 120), 180);
})();
</script>
                """,
                height=0,
            )
            st.success("Audio armed.")

        st.subheader("Maintenance")
        if st.button("Clear EVE Memory (SQLite)"):
            clear_msgs()
            st.session_state.history = []
            st.success("Memory cleared.")
            st.rerun()

    # Bottom Dock (never blocks content: it collapses)
    st.components.v1.html(dock_html(wake), height=120, scrolling=False)


# -----------------------------
# Entry
# -----------------------------
if __name__ == "__main__":
    st.set_page_config(page_title=APP_TITLE, layout="wide", initial_sidebar_state="collapsed")
    main()


# /requirements.txt
# streamlit==1.53.1
# pandas==2.3.3
# openpyxl==3.1.5
# plotly==6.5.2
# requests==2.32.5
# reportlab==4.4.4
# openai>=1.0.0
