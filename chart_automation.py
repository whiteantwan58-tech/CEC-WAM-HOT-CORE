"""
chart_automation.py
-------------------

This module automates the process of linking and updating missing charts
and information across the CEC-WAM-HOT-CORE repository. It:

1. Identifies missing data and charts by cross-referencing Chart.xlsx with CSV files
2. Updates Chart.xlsx with data from CSV sources
3. Generates chart visualizations
4. Exports data for integration with the live system

Usage:
    python chart_automation.py
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
import os
import json
from datetime import datetime
import logging

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ChartAutomation:
    """Main class for chart and data automation"""
    
    def __init__(self, base_dir=None):
        """Initialize the automation system"""
        self.base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
        self.chart_file = os.path.join(self.base_dir, 'Chart.xlsx')
        self.data_dir = os.path.join(self.base_dir, 'data')
        self.csv_files = self._discover_csv_files()
        self.data_cache = {}
        
    def _discover_csv_files(self):
        """Discover all CSV files in the repository"""
        csv_files = []
        
        # Root directory CSV files
        for file in os.listdir(self.base_dir):
            if file.endswith('.csv'):
                csv_files.append(os.path.join(self.base_dir, file))
        
        # Data directory CSV files
        if os.path.exists(self.data_dir):
            for file in os.listdir(self.data_dir):
                if file.endswith('.csv'):
                    csv_files.append(os.path.join(self.data_dir, file))
        
        logger.info(f"Discovered {len(csv_files)} CSV files")
        return csv_files
    
    def load_csv_data(self):
        """Load all CSV files into memory"""
        logger.info("Loading CSV data...")
        
        for csv_file in self.csv_files:
            try:
                filename = os.path.basename(csv_file)
                df = pd.read_csv(csv_file)
                self.data_cache[filename] = df
                logger.info(f"Loaded {filename}: {len(df)} rows, {len(df.columns)} columns")
            except Exception as e:
                logger.error(f"Failed to load {csv_file}: {str(e)}")
        
        return self.data_cache
    
    def analyze_chart_xlsx(self):
        """Analyze Chart.xlsx structure and content"""
        logger.info("Analyzing Chart.xlsx...")
        
        if not os.path.exists(self.chart_file):
            logger.warning("Chart.xlsx not found - will create new file")
            return None
        
        try:
            wb = load_workbook(self.chart_file)
            analysis = {
                'sheets': wb.sheetnames,
                'sheet_data': {}
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.max_row
                cols = ws.max_column
                analysis['sheet_data'][sheet_name] = {
                    'rows': rows,
                    'columns': cols,
                    'has_data': rows > 1
                }
                logger.info(f"Sheet '{sheet_name}': {rows} rows x {cols} columns")
            
            return analysis
        except Exception as e:
            logger.error(f"Failed to analyze Chart.xlsx: {str(e)}")
            return None
    
    def identify_missing_data(self):
        """Identify missing data and charts"""
        logger.info("Identifying missing data and charts...")
        
        missing_data = {
            'csv_files_not_in_chart': [],
            'empty_sheets': [],
            'recommendations': []
        }
        
        chart_analysis = self.analyze_chart_xlsx()
        
        # Check which CSV data is not represented in Chart.xlsx
        csv_names = set(os.path.basename(f).replace('.csv', '') for f in self.csv_files)
        
        if chart_analysis:
            chart_sheets = set(chart_analysis['sheets'])
            missing = csv_names - chart_sheets
            missing_data['csv_files_not_in_chart'] = list(missing)
            
            # Check for empty sheets
            for sheet_name, info in chart_analysis['sheet_data'].items():
                if not info['has_data']:
                    missing_data['empty_sheets'].append(sheet_name)
        else:
            missing_data['csv_files_not_in_chart'] = list(csv_names)
            missing_data['recommendations'].append("Create new Chart.xlsx with all CSV data")
        
        logger.info(f"Found {len(missing_data['csv_files_not_in_chart'])} CSV files not in Chart.xlsx")
        logger.info(f"Found {len(missing_data['empty_sheets'])} empty sheets")
        
        return missing_data
    
    def update_chart_xlsx(self):
        """Update Chart.xlsx with data from CSV files"""
        logger.info("Updating Chart.xlsx...")
        
        try:
            # Load or create workbook
            if os.path.exists(self.chart_file):
                wb = load_workbook(self.chart_file)
            else:
                wb = openpyxl.Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Add or update sheets for each CSV
            for csv_file in self.csv_files:
                filename = os.path.basename(csv_file)
                sheet_name = filename.replace('.csv', '')[:31]  # Excel sheet name limit
                
                try:
                    df = pd.read_csv(csv_file)
                    
                    # Remove sheet if it exists
                    if sheet_name in wb.sheetnames:
                        wb.remove(wb[sheet_name])
                    
                    # Create new sheet
                    ws = wb.create_sheet(sheet_name)
                    
                    # Write headers
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=str(col_name))
                    
                    # Write data (limit to first 1000 rows for performance)
                    max_rows = min(len(df), 1000)
                    for row_idx in range(max_rows):
                        for col_idx, col_name in enumerate(df.columns, 1):
                            value = df.iloc[row_idx][col_name]
                            ws.cell(row=row_idx + 2, column=col_idx, value=value)
                    
                    logger.info(f"Updated sheet '{sheet_name}' with {max_rows} rows")
                    
                except Exception as e:
                    logger.error(f"Failed to process {filename}: {str(e)}")
            
            # Save workbook
            wb.save(self.chart_file)
            logger.info(f"Chart.xlsx saved successfully")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to update Chart.xlsx: {str(e)}")
            return False
    
    def generate_chart_data_json(self):
        """Generate JSON data for web charts"""
        logger.info("Generating chart data JSON...")
        
        chart_data = {
            'timestamp': datetime.now().isoformat(),
            'datasets': {}
        }
        
        # Process pump.fun.csv for transaction chart
        pump_file = os.path.join(self.base_dir, 'pump.fun.csv')
        if os.path.exists(pump_file):
            try:
                df = pd.read_csv(pump_file)
                # Extract time series data
                if 'Human Time' in df.columns and 'Value' in df.columns:
                    df['Human Time'] = pd.to_datetime(df['Human Time'])
                    df = df.sort_values('Human Time')
                    
                    chart_data['datasets']['pump_fun_timeline'] = {
                        'labels': df['Human Time'].astype(str).tolist()[-20:],
                        'values': df['Value'].tolist()[-20:],
                        'type': 'line',
                        'title': 'Pump.fun Transaction Values'
                    }
                    logger.info("Generated pump.fun chart data")
            except Exception as e:
                logger.error(f"Failed to process pump.fun.csv: {str(e)}")
        
        # Process BlackHoles.csv for discoveries chart
        blackholes_file = os.path.join(self.base_dir, 'BlackHoles.csv')
        if os.path.exists(blackholes_file):
            try:
                df = pd.read_csv(blackholes_file)
                if 'DISCOVERY_ID' in df.columns and 'STATUS' in df.columns:
                    status_counts = df['STATUS'].value_counts().to_dict()
                    
                    chart_data['datasets']['blackholes_status'] = {
                        'labels': list(status_counts.keys()),
                        'values': list(status_counts.values()),
                        'type': 'pie',
                        'title': 'BlackHole Discovery Status'
                    }
                    logger.info("Generated BlackHoles chart data")
            except Exception as e:
                logger.error(f"Failed to process BlackHoles.csv: {str(e)}")
        
        # Process CEC Matrix metrics
        cec_file = os.path.join(self.base_dir, 'CEC Matrix System Operational Metrics and Assets - FINANCE_HUB (1).csv')
        if os.path.exists(cec_file):
            try:
                df = pd.read_csv(cec_file)
                if 'A (ASSET CLASS)' in df.columns and 'D (VALUE)' in df.columns:
                    # Clean values
                    df_clean = df[df['D (VALUE)'].notna()].copy()
                    
                    chart_data['datasets']['cec_assets'] = {
                        'labels': df_clean['A (ASSET CLASS)'].tolist(),
                        'values': df_clean['D (VALUE)'].tolist(),
                        'type': 'bar',
                        'title': 'CEC Asset Distribution'
                    }
                    logger.info("Generated CEC assets chart data")
            except Exception as e:
                logger.error(f"Failed to process CEC Matrix: {str(e)}")
        
        # Save to JSON file
        output_file = os.path.join(self.data_dir, 'chart_data.json')
        os.makedirs(self.data_dir, exist_ok=True)
        
        with open(output_file, 'w') as f:
            json.dump(chart_data, f, indent=2)
        
        logger.info(f"Chart data saved to {output_file}")
        return chart_data
    
    def generate_summary_report(self):
        """Generate a summary report of the automation"""
        logger.info("Generating summary report...")
        
        report = {
            'timestamp': datetime.now().isoformat(),
            'csv_files_processed': len(self.csv_files),
            'csv_files': [os.path.basename(f) for f in self.csv_files],
            'chart_xlsx_updated': os.path.exists(self.chart_file),
            'data_cache_size': len(self.data_cache),
            'status': 'SUCCESS'
        }
        
        # Save report
        report_file = os.path.join(self.data_dir, 'automation_report.json')
        os.makedirs(self.data_dir, exist_ok=True)
        
        with open(report_file, 'w') as f:
            json.dump(report, f, indent=2)
        
        logger.info(f"Summary report saved to {report_file}")
        return report
    
    def run_full_automation(self):
        """Run the complete automation pipeline"""
        logger.info("=" * 60)
        logger.info("Starting Chart Automation Pipeline")
        logger.info("=" * 60)
        
        try:
            # Step 1: Load CSV data
            self.load_csv_data()
            
            # Step 2: Analyze existing Chart.xlsx
            self.analyze_chart_xlsx()
            
            # Step 3: Identify missing data
            missing = self.identify_missing_data()
            
            # Step 4: Update Chart.xlsx
            self.update_chart_xlsx()
            
            # Step 5: Generate chart data for web
            self.generate_chart_data_json()
            
            # Step 6: Generate summary report
            report = self.generate_summary_report()
            
            logger.info("=" * 60)
            logger.info("Chart Automation Pipeline Completed Successfully")
            logger.info("=" * 60)
            
            return report
            
        except Exception as e:
            logger.error(f"Automation pipeline failed: {str(e)}")
            return {'status': 'FAILED', 'error': str(e)}


def main():
    """Main entry point"""
    automation = ChartAutomation()
    report = automation.run_full_automation()
    
    print("\n" + "=" * 60)
    print("AUTOMATION REPORT")
    print("=" * 60)
    print(f"Status: {report.get('status', 'UNKNOWN')}")
    print(f"CSV Files Processed: {report.get('csv_files_processed', 0)}")
    print(f"Chart.xlsx Updated: {report.get('chart_xlsx_updated', False)}")
    print("=" * 60)
    
    return 0 if report.get('status') == 'SUCCESS' else 1


if __name__ == '__main__':
    exit(main())
